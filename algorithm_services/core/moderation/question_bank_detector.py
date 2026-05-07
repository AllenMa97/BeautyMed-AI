"""题库检测器 - 按C35-C39分类的语义相似度检测"""
import asyncio
import openpyxl
import json
import os
import threading
import time
from typing import Dict, List, Set, Optional, Any
from dataclasses import dataclass
from concurrent.futures import ThreadPoolExecutor, as_completed
import numpy as np
import jieba
from dotenv import load_dotenv
load_dotenv(os.path.join(os.path.dirname(__file__), "../../config/MODERATION.env"))

from algorithm_services.utils.logger import get_logger
from algorithm_services.large_model.llm_factory import get_embedding_client, EmbeddingRequest
from algorithm_services.core.moderation.embedding_storage import get_embedding_storage


logger = get_logger(__name__)


QUESTION_BANK_C_THRESHOLDS = {
    "C35": float(os.getenv("QUESTION_BANK_THRESHOLD_C35", "0.85")),
    "C36": float(os.getenv("QUESTION_BANK_THRESHOLD_C36", "0.85")),
    "C37": float(os.getenv("QUESTION_BANK_THRESHOLD_C37", "0.85")),
    "C38": float(os.getenv("QUESTION_BANK_THRESHOLD_C38", "0.85")),
    "C39": float(os.getenv("QUESTION_BANK_THRESHOLD_C39", "0.85")),
}


QUESTION_BANK_C_CODES = {
    "生成内容测试题集（人工+关键词抽检）": "C35",
    "拒答题": "C36",
    "涉知识产权、商业秘密评估": "C37",
    "涉准确性、可靠性评估": "C38",
    "涉民族、信仰、性别等评估": "C39",
}


@dataclass
class QuestionData:
    """问题数据结构"""
    id: int
    question: str
    risk_subclass: str
    c_code: str
    embedding: Optional[List[float]] = None


class QuestionBankDetector:
    """题库检测器 - 按C35-C39分类的语义相似度并行检测"""

    def __init__(self, excel_path: str = r"D:\lansee_chatbot\algorithm_services\data\附件（5）评估测试题.xlsx"):
        self.excel_path = excel_path
        self.question_data: Dict[str, List[QuestionData]] = {}
        self.all_questions: List[QuestionData] = []
        self.question_embeddings: Dict[str, np.ndarray] = {}
        self.json_storage_path = "data/question_bank_json"
        self.embedding_storage = None
        self._precompute_done = False
        self._embeddings_lock = threading.Lock()
        self._init_embedding_storage()
        self._load_question_data()
        precompute_thread = threading.Thread(target=self._precompute_embeddings, daemon=True)
        precompute_thread.start()

    def _init_embedding_storage(self):
        """初始化embedding存储"""
        try:
            self.embedding_storage = get_embedding_storage()
            logger.info("[题库检测器] Embedding存储初始化成功")
        except Exception as e:
            logger.warning(f"[题库检测器] Embedding存储初始化失败: {e}")
            self.embedding_storage = None

    def _load_question_data(self):
        """加载题库数据：优先从JSON读取，JSON不存在时从Excel导出"""
        os.makedirs(self.json_storage_path, exist_ok=True)

        json_files = {
            "C35": os.path.join(self.json_storage_path, "C35.json"),
            "C36": os.path.join(self.json_storage_path, "C36.json"),
            "C37": os.path.join(self.json_storage_path, "C37.json"),
            "C38": os.path.join(self.json_storage_path, "C38.json"),
            "C39": os.path.join(self.json_storage_path, "C39.json"),
        }

        all_exist = all(os.path.exists(f) for f in json_files.values())

        if all_exist:
            self._load_from_json(json_files)
        else:
            self._load_excel_data()
            self._save_to_json()

    def _load_from_json(self, json_files: Dict[str, str]):
        """从JSON文件加载题库数据"""
        logger.info("[题库检测器] 从JSON缓存加载题库数据...")

        total_questions = 0
        for c_code, filepath in json_files.items():
            try:
                with open(filepath, 'r', encoding='utf-8') as f:
                    data = json.load(f)

                questions = []
                for item in data:
                    questions.append(QuestionData(
                        id=item.get('id', 0),
                        question=item['question'],
                        risk_subclass=item.get('risk_subclass', ''),
                        c_code=c_code
                    ))

                self.question_data[c_code] = questions
                self.all_questions.extend(questions)
                total_questions += len(questions)

            except Exception as e:
                logger.warning(f"[题库检测器] 加载 {filepath} 失败: {e}，回退到Excel加载")
                self._load_excel_data()
                self._save_to_json()
                return

        logger.info(f"[题库检测器] JSON缓存加载完成，总违规问题数: {total_questions}")

    def _load_excel_data(self):
        """加载Excel数据到内存，按C35-C39分类"""
        logger.info(f"[题库检测器] 开始加载Excel数据: {self.excel_path}")
        wb = openpyxl.load_workbook(self.excel_path, read_only=True)

        total_questions = 0
        for sheet_name in QUESTION_BANK_C_CODES.keys():
            if sheet_name not in wb.sheetnames:
                logger.warning(f"[题库检测器] Sheet不存在: {sheet_name}")
                continue

            c_code = QUESTION_BANK_C_CODES[sheet_name]
            ws = wb[sheet_name]
            questions = []

            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row[1] or not str(row[1]).strip():
                    continue

                question_data = self._parse_row(sheet_name, c_code, row)
                if question_data:
                    questions.append(question_data)
                    total_questions += 1

            self.question_data[c_code] = questions
            self.all_questions.extend(questions)

            logger.info(f"[题库检测器] 加载Sheet: {sheet_name} ({c_code}), 违规问题数: {len(questions)}")

        wb.close()
        logger.info(f"[题库检测器] Excel数据加载完成，总违规问题数: {total_questions}")

    def _parse_row(self, sheet_name: str, c_code: str, row: tuple) -> Optional[QuestionData]:
        """根据不同的sheet结构解析行数据"""
        try:
            if sheet_name == '生成内容测试题集（人工+关键词抽检）':
                return QuestionData(
                    id=int(row[0]) if row[0] else 0,
                    question=str(row[1]).strip(),
                    risk_subclass=str(row[2]).strip() if row[2] else "",
                    c_code=c_code
                )
            elif sheet_name == '拒答题':
                return QuestionData(
                    id=int(row[0]) if row[0] else 0,
                    question=str(row[1]).strip(),
                    risk_subclass=str(row[2]).strip() if row[2] else "",
                    c_code=c_code
                )
            elif sheet_name == '涉知识产权、商业秘密评估':
                return QuestionData(
                    id=int(row[1]) if row[1] else 0,
                    question=str(row[2]).strip(),
                    risk_subclass=str(row[3]).strip() if row[3] else "",
                    c_code=c_code
                )
            elif sheet_name in ['涉准确性、可靠性评估', '涉民族、信仰、性别等评估']:
                return QuestionData(
                    id=int(row[0]) if row[0] else 0,
                    question=str(row[1]).strip(),
                    risk_subclass=str(row[2]).strip() if row[2] else "",
                    c_code=c_code
                )
            else:
                return None
        except Exception as e:
            logger.warning(f"[题库检测器] 解析行数据失败: {e}, sheet={sheet_name}")
            return None
    
    def _save_to_json(self):
        """保存到JSON文件（只保存必要信息）"""
        os.makedirs(self.json_storage_path, exist_ok=True)
        
        for sheet_name, questions in self.question_data.items():
            json_data = []
            for q in questions:
                question_json = {
                    'id': q.id,
                    'question': q.question,
                    'risk_subclass': q.risk_subclass
                }
                json_data.append(question_json)
            
            file_path = os.path.join(self.json_storage_path, f"{sheet_name}.json")
            with open(file_path, 'w', encoding='utf-8') as f:
                json.dump(json_data, f, ensure_ascii=False, indent=2)
            
            logger.info(f"[题库检测器] 保存JSON: {file_path}, 问题数: {len(json_data)}")
    
    async def _get_embedding(self, text: str) -> List[float]:
        """获取文本的embedding"""
        try:
            embedding_client = get_embedding_client()
            
            request = EmbeddingRequest(
                text=text,
                model="text-embedding-v3",
                dimensions=1024
            )
            
            embedding = await embedding_client.get_embedding(request)
            
            if embedding:
                return embedding
            
            logger.warning(f"[题库检测器] 获取embedding失败: {text[:50]}...")
            return []
            
        except Exception as e:
            logger.warning(f"[题库检测器] 获取embedding异常: {e}")
            return []
    
    def _precompute_embeddings(self):
        """预计算所有问题的embedding（按C码分组并行处理，批量获取，支持持久化）"""
        logger.info(f"[题库检测器] 开始预计算embedding，共 {len(self.all_questions)} 个问题...")

        if not self.embedding_storage:
            logger.warning("[题库检测器] Embedding存储未初始化，跳过持久化")

        BATCH_SIZE = 6
        MAX_RETRIES = 3
        RETRY_DELAY = 2

        def compute_embedding_for_c_code(c_code, questions):
            """为单个C码批量计算embeddings"""
            sheet_results = []
            need_compute = []
            need_compute_questions = []

            for question in questions:
                existing_embedding = None
                if self.embedding_storage:
                    existing_embedding = self.embedding_storage.get_embedding(question.question)

                if existing_embedding:
                    sheet_results.append((question.question, np.array(existing_embedding)))
                else:
                    need_compute.append(question.question)
                    need_compute_questions.append(question)

            if need_compute:
                logger.info(f"[题库检测器] {c_code}: 缓存命中 {len(sheet_results)} 个, 需计算 {len(need_compute)} 个")

                for batch_start in range(0, len(need_compute), BATCH_SIZE):
                    batch_end = min(batch_start + BATCH_SIZE, len(need_compute))
                    batch_texts = need_compute[batch_start:batch_end]
                    batch_questions = need_compute_questions[batch_start:batch_end]
                    batch_num = batch_start // BATCH_SIZE + 1

                    for attempt in range(MAX_RETRIES):
                        loop = asyncio.new_event_loop()
                        asyncio.set_event_loop(loop)
                        try:
                            embedding_client = get_embedding_client()
                            embeddings = loop.run_until_complete(
                                embedding_client.get_embeddings_batch(batch_texts, model="text-embedding-v3", dimensions=1024)
                            )

                            for i, embedding in enumerate(embeddings):
                                if embedding:
                                    embedding_array = np.array(embedding)
                                    sheet_results.append((batch_questions[i].question, embedding_array))

                                    if self.embedding_storage:
                                        self.embedding_storage.add_embedding(
                                            text=batch_questions[i].question,
                                            embedding=embedding,
                                            model_name="text-embedding-v3",
                                            source=f"question_bank_{c_code}",
                                            extra_info={'risk_subclass': batch_questions[i].risk_subclass}
                                        )

                            logger.info(f"[题库检测器] {c_code}: 批次 {batch_num} 完成 ({batch_start+1}-{batch_end}/{len(need_compute)})")
                            break

                        except Exception as e:
                            if attempt < MAX_RETRIES - 1:
                                logger.warning(f"[题库检测器] {c_code} 批次 {batch_num} 第{attempt+1}次失败: {e}, {RETRY_DELAY}秒后重试...")
                                time.sleep(RETRY_DELAY)
                            else:
                                logger.warning(f"[题库检测器] {c_code} 批次 {batch_num} 计算失败（已重试{MAX_RETRIES}次）: {e}")
                        finally:
                            loop.close()

                    if self.embedding_storage:
                        self.embedding_storage.flush()

            return sheet_results

        with ThreadPoolExecutor(max_workers=5) as executor:
            futures = {
                executor.submit(compute_embedding_for_c_code, c_code, questions): c_code
                for c_code, questions in self.question_data.items()
            }

            for future in as_completed(futures):
                c_code = futures[future]
                try:
                    sheet_results = future.result()
                    with self._embeddings_lock:
                        for question_text, embedding in sheet_results:
                            self.question_embeddings[question_text] = embedding
                    logger.info(f"[题库检测器] {c_code} embedding计算完成")
                except Exception as e:
                    logger.warning(f"[题库检测器] {c_code} embedding计算失败: {e}")

        logger.info(f"[题库检测器] embedding预计算完成，成功 {len(self.question_embeddings)} 个问题")
        self._precompute_done = True
    
    def _calculate_similarity(self, text1: str, text2: str) -> float:
        """计算两个文本的相似度（余弦相似度）"""
        if text1 not in self.question_embeddings or text2 not in self.question_embeddings:
            return 0.0
        
        embedding1 = self.question_embeddings[text1]
        embedding2 = self.question_embeddings[text2]
        
        dot_product = np.dot(embedding1, embedding2)
        norm1 = np.linalg.norm(embedding1)
        norm2 = np.linalg.norm(embedding2)
        
        if norm1 == 0 or norm2 == 0:
            return 0.0
        
        cosine_similarity = dot_product / (norm1 * norm2)
        return float(cosine_similarity)
    
    async def detect_semantic_violation(self, text: str, threshold: float = None) -> Optional[Dict[str, Any]]:
        """基于语义相似度检测违规问题（按sheet分组并行处理，支持持久化）"""
        logger.info(f"[题库检测器] 开始语义检测: {text[:50]}..., threshold={threshold}")

        if not self.question_embeddings:
            if not self._precompute_done:
                logger.info(f"[题库检测器] embedding预计算尚未完成且无可用缓存，跳过语义检测")
            else:
                logger.info(f"[题库检测器] 无可用embedding，跳过语义检测")
            return None
        
        text_embedding = await self._get_embedding(text)
        if not text_embedding:
            logger.warning(f"[题库检测器] 获取用户输入embedding失败")
            return None
        
        text_embedding_array = np.array(text_embedding)

        c_code_results = {}
        
        if self.embedding_storage:
            self.embedding_storage.add_embedding(
                text=text,
                embedding=text_embedding,
                model_name="text-embedding-v3",
                source="user_query",
                extra_info={'detection_threshold': threshold}
            )
        
        def compute_similarity_for_c_code(c_code, questions):
            """为单个C码计算相似度，直接矩阵运算"""
            c_code_max_similarity = 0.0
            c_code_best_match = None

            question_texts = [q.question for q in questions]
            available_questions = [q for q in question_texts if q in self.question_embeddings]

            if not available_questions:
                return c_code, c_code_max_similarity, c_code_best_match

            question_embeddings = np.array([self.question_embeddings[q] for q in available_questions])

            similarities = np.dot(text_embedding_array, question_embeddings.T)

            max_idx = np.argmax(similarities)
            c_code_max_similarity = float(similarities[max_idx])

            for i, q in enumerate(available_questions):
                if i == max_idx:
                    for q_obj in questions:
                        if q_obj.question == q:
                            c_code_best_match = q_obj
                            break
                    break

            return c_code, c_code_max_similarity, c_code_best_match

        max_similarity = 0.0
        best_match = None
        best_c_code = None
        c_code_thresholds = {}

        for c_code in self.question_data.keys():
            c_code_thresholds[c_code] = QUESTION_BANK_C_THRESHOLDS.get(c_code, 0.85)

        with ThreadPoolExecutor(max_workers=5) as executor:
            futures = {
                executor.submit(compute_similarity_for_c_code, c_code, questions): c_code
                for c_code, questions in self.question_data.items()
            }

            for future in as_completed(futures):
                c_code = futures[future]
                try:
                    c_code, sheet_max_similarity, sheet_best_match = future.result()

                    if sheet_best_match and sheet_max_similarity >= c_code_thresholds[c_code]:
                        if sheet_max_similarity > max_similarity:
                            max_similarity = sheet_max_similarity
                            best_match = sheet_best_match
                            best_c_code = c_code

                except Exception as e:
                    logger.warning(f"[题库检测器] C码 {c_code} 相似度计算失败: {e}")

        if max_similarity > 0 and best_match and best_c_code:
            logger.warning(f"[题库检测器] {best_c_code} 检测到语义违规: 相似度={max_similarity:.3f}, 问题={best_match.question[:50]}...")
            return {
                'is_violation': True,
                'similarity': max_similarity,
                'matched_question': best_match.question,
                'risk_subclass': best_match.risk_subclass,
                'c_code': best_c_code,
                'threshold': c_code_thresholds[best_c_code]
            }

        logger.info(f"[题库检测器] 语义检测通过: {text[:50]}...")
        return {
            'is_violation': False,
            'similarity': max_similarity,
            'matched_question': None,
            'threshold': threshold
        }
    
    def _find_sheet_for_question(self, question: QuestionData) -> str:
        """找到问题所属的sheet"""
        for sheet_name, questions in self.question_data.items():
            if question in questions:
                return sheet_name
        return "unknown"
    
    def get_statistics(self) -> Dict[str, Any]:
        """获取统计信息"""
        stats = {
            'total_sheets': len(self.question_data),
            'total_questions': len(self.all_questions),
            'questions_per_sheet': {
                sheet_name: len(questions)
                for sheet_name, questions in self.question_data.items()
            },
            'embeddings_computed': len(self.question_embeddings),
            'json_files_saved': len(self.question_data)
        }
        
        if self.embedding_storage:
            storage_stats = self.embedding_storage.get_statistics()
            stats['embedding_storage'] = storage_stats
        
        return stats
    
    async def add_new_question(self, question: str, risk_subclass: str, 
                       sheet_name: str = "新增问题", embedding: Optional[List[float]] = None) -> bool:
        """动态添加新问题到题库"""
        try:
            new_id = len(self.all_questions) + 1
            
            if embedding is None:
                embedding = await self._get_embedding(question)
                if not embedding:
                    logger.warning(f"[题库检测器] 获取新问题embedding失败: {query[:50]}...")
                    return False
            
            question_data = QuestionData(
                id=new_id,
                question=question,
                risk_subclass=risk_subclass,
                embedding=embedding
            )
            
            if sheet_name not in self.question_data:
                self.question_data[sheet_name] = []
            
            self.question_data[sheet_name].append(question_data)
            self.all_questions.append(question_data)
            self.question_embeddings[question] = np.array(embedding)
            
            if self.embedding_storage:
                self.embedding_storage.add_embedding(
                    text=question,
                    embedding=embedding,
                    model_name="text-embedding-v3",
                    source=f"question_bank_{sheet_name}",
                    extra_info={'risk_subclass': risk_subclass, 'is_new': True}
                )
            
            self._save_to_json()
            
            logger.info(f"[题库检测器] 添加新问题到题库: {question[:50]}..., sheet={sheet_name}")
            return True
            
        except Exception as e:
            logger.warning(f"[题库检测器] 添加新问题失败: {e}")
            return False
    
    def batch_add_questions(self, questions_data: List[Dict[str, Any]], 
                         sheet_name: str = "新增问题") -> int:
        """批量添加新问题到题库"""
        added_count = 0
        
        for data in questions_data:
            question = data.get('query', '')
            risk_subclass = data.get('risk_subclass', '未知')
            embedding = data.get('embedding')
            
            if self.add_new_question(question, risk_subclass, sheet_name, embedding):
                added_count += 1
        
        logger.info(f"[题库检测器] 批量添加新问题完成: {added_count}/{len(questions_data)}")
        return added_count
    
    def merge_candidates_to_question_bank(self, candidates_file: str = "data/blocked_queries/question_bank_candidates.json",
                                     min_similarity: float = 0.9, min_frequency: int = 5) -> int:
        """将候选问题合并到题库"""
        if not os.path.exists(candidates_file):
            logger.warning(f"[题库检测器] 候选文件不存在: {candidates_file}")
            return 0
        
        try:
            with open(candidates_file, 'r', encoding='utf-8') as f:
                candidates = json.load(f)
            
            approved_candidates = []
            for candidate in candidates:
                if candidate.get('approved', False) and candidate.get('review_count', 0) >= min_frequency:
                    approved_candidates.append(candidate)
            
            if not approved_candidates:
                logger.info(f"[题库检测器] 没有符合条件的候选问题")
                return 0
            
            added_count = 0
            for candidate in approved_candidates:
                similarity = candidate.get('similarity', 0)
                if similarity >= min_similarity:
                    risk_subclass = candidate.get('risk_subclass', '未知')
                    embedding = candidate.get('embedding')
                    
                    if self.add_new_question(candidate['query'], risk_subclass, "新增问题", embedding):
                        added_count += 1
            
            logger.info(f"[题库检测器] 合并候选问题到题库: {added_count}/{len(approved_candidates)}")
            return added_count
            
        except Exception as e:
            logger.warning(f"[题库检测器] 合并候选问题失败: {e}")
            return 0


_question_bank_detector = None


def get_question_bank_detector() -> QuestionBankDetector:
    """获取题库检测器单例"""
    global _question_bank_detector
    if _question_bank_detector is None:
        _question_bank_detector = QuestionBankDetector()
    return _question_bank_detector
