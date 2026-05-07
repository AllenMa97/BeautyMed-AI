"""Excel词表检测器 - 按C8-C34风险子类分类的并行检测"""
import asyncio
import openpyxl
from typing import Dict, List, Set, Tuple, Optional
from dataclasses import dataclass
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path
import jieba
from algorithm_services.utils.logger import get_logger


logger = get_logger(__name__)


RISK_SUBCLASS_CODES = {
    "A.1 a) 煽动颠覆国家政权、推翻社会主义制度": "C8",
    "A.1 b) 危害国家安全和利益、损害国家形象": "C9",
    "A.1 c) 煽动分裂国家、破坏国家统一和社会稳定": "C10",
    "A.1 d) 宣扬恐怖主义、极端主义": "C11",
    "A.1 e) 宣扬民族仇恨": "C12",
    "A.1 f) 宣扬暴力、淫秽色情": "C13",
    "A.1 g) 传播虚假有害信息": "C14",
    "A.1 h) 其他法律、行政法规禁止的内容": "C15",
    "A.2 a) 民族歧视内容": "C16",
    "A.2 b) 信仰歧视内容": "C17",
    "A.2 c) 国别歧视内容": "C18",
    "A.2 d) 地域歧视内容": "C19",
    "A.2 e) 性别歧视内容": "C20",
    "A.2 f) 年龄歧视内容": "C21",
    "A.2 g) 职业歧视内容": "C22",
    "A.2 h) 健康歧视内容": "C23",
    "A.2 i) 其他方面歧视内容": "C24",
    "A.3 a) 侵犯他人知识产权": "C25",
    "A.3 b) 违反商业道德": "C26",
    "A.3 c) 泄露他人商业秘密": "C27",
    "A.3 d) 利用算法、数据、平台等优势，实施垄断和不正当竞争行为": "C28",
    "A.4 a) 危害他人身心健康": "C29",
    "A.4 b) 侵害他人肖像权": "C30",
    "A.4 c) 侵害他人名誉权": "C31",
    "A.4 d) 侵害他人荣誉权": "C32",
    "A.4 e) 侵害他人隐私权": "C33",
    "A.4 f) 侵害他人个人信息权益": "C34",
}


@dataclass
class ViolationResult:
    """违规检测结果"""
    is_violation: bool
    c_code: str
    risk_subclass: str
    keyword: str


class ParallelKeywordDetector:
    """Excel词表检测器 - 按C8-C34风险子类并行检测"""

    def __init__(self, excel_path: str = None):
        if excel_path is None:
            excel_path = str(Path(__file__).resolve().parent.parent.parent / "data" / "附件（4）拦截关键词列表.xlsx")
        self.excel_path = excel_path
        self.risk_subclass_keywords: Dict[str, List[Tuple[str, str]]] = {}
        self.risk_subclass_sets: Dict[str, Set[str]] = {}
        self.risk_subclass_names: Dict[str, str] = {}
        self._load_excel_data()

    def _load_excel_data(self):
        """加载Excel数据到内存，按C8-C34风险子类分类"""
        try:
            logger.info(f"[Excel词表检测] 开始加载Excel数据: {self.excel_path}")
            wb = openpyxl.load_workbook(self.excel_path, read_only=True)

            total_keywords = 0
            filtered_count = 0
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]

                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row[0] and str(row[0]).strip():
                        keyword = str(row[0]).strip()
                        risk_subclass = str(row[1]).strip() if row[1] else ""

                        # 过滤掉长度小于2的关键词，避免误报（如单独的字母"b"）
                        if len(keyword) < 2:
                            filtered_count += 1
                            continue

                        c_code = RISK_SUBCLASS_CODES.get(risk_subclass, None)
                        if not c_code:
                            continue

                        if c_code not in self.risk_subclass_keywords:
                            self.risk_subclass_keywords[c_code] = []
                            self.risk_subclass_sets[c_code] = set()
                            self.risk_subclass_names[c_code] = risk_subclass

                        self.risk_subclass_keywords[c_code].append((keyword, risk_subclass))
                        self.risk_subclass_sets[c_code].add(keyword)
                        total_keywords += 1

            wb.close()
            logger.info(f"[Excel词表检测] Excel数据加载完成，总关键词数: {total_keywords}, 过滤短词数: {filtered_count}, C码数量: {len(self.risk_subclass_keywords)}")

        except Exception as e:
            logger.error(f"[Excel词表检测] 加载Excel数据失败: {e}")
            raise

    def _segment_text(self, text: str) -> Set[str]:
        """使用jieba分词"""
        try:
            words = jieba.lcut(text)
            return set(words)
        except Exception as e:
            logger.warning(f"[Excel词表检测] jieba分词失败: {e}")
            return set([text])

    def _check_single_c_code(self, text: str, c_code: str, chunk_size: int = 50) -> Optional[ViolationResult]:
        """检查单个C码的关键词，支持chunk并行"""
        keywords = self.risk_subclass_keywords.get(c_code, [])
        if not keywords:
            return None

        text_lower = text.lower()

        if len(keywords) <= chunk_size:
            for keyword, risk_subclass in keywords:
                if keyword.lower() in text_lower:
                    return ViolationResult(
                        is_violation=True,
                        c_code=c_code,
                        risk_subclass=risk_subclass,
                        keyword=keyword
                    )
            return None

        chunks = [keywords[i:i + chunk_size] for i in range(0, len(keywords), chunk_size)]

        def check_chunk(chunk):
            for keyword, risk_subclass in chunk:
                if keyword.lower() in text_lower:
                    return ViolationResult(
                        is_violation=True,
                        c_code=c_code,
                        risk_subclass=risk_subclass,
                        keyword=keyword
                    )
            return None

        with ThreadPoolExecutor(max_workers=len(chunks)) as chunk_executor:
            futures = {chunk_executor.submit(check_chunk, chunk): chunk for chunk in chunks}
            for future in as_completed(futures):
                result = future.result()
                if result and result.is_violation:
                    return result

        return None

    def detect_parallel(self, text: str) -> ViolationResult:
        """C8-C34 全部并行检测，每个C码一个线程"""
        logger.info(f"[Excel词表检测-并行] 开始检测文本: {text[:50]}...")

        c_codes = list(self.risk_subclass_keywords.keys())

        with ThreadPoolExecutor(max_workers=len(c_codes)) as executor:
            futures = {
                executor.submit(self._check_single_c_code, text, c_code): c_code
                for c_code in c_codes
            }

            for future in as_completed(futures):
                result = future.result()
                if result and result.is_violation:
                    logger.warning(f"[Excel词表检测-并行] {result.c_code} 检测到违规: {result.keyword}")
                    return result

        logger.info(f"[Excel词表检测-并行] 文本通过检测: {text[:50]}...")
        return ViolationResult(
            is_violation=False,
            c_code="",
            risk_subclass="",
            keyword=""
        )

    def detect_parallel_with_segmentation(self, text: str) -> ViolationResult:
        """C8-C34 全部并行检测 + 分词匹配辅助 + chunk并行"""
        logger.info(f"[Excel词表检测-并行+分词] 开始检测文本: {text[:50]}...")

        text_lower = text.lower()
        words = self._segment_text(text)
        c_codes = list(self.risk_subclass_keywords.keys())

        def check_with_both(c_code: str) -> Optional[ViolationResult]:
            keywords = self.risk_subclass_keywords.get(c_code, [])
            if not keywords:
                return None

            keyword_set = self.risk_subclass_sets.get(c_code, set())

            direct_match = self._check_single_c_code(text, c_code, chunk_size=50)
            if direct_match:
                return direct_match

            intersection = words & keyword_set
            if intersection:
                keyword = list(intersection)[0]
                risk_subclass = self.risk_subclass_names.get(c_code, "")
                return ViolationResult(
                    is_violation=True,
                    c_code=c_code,
                    risk_subclass=risk_subclass,
                    keyword=keyword
                )

            return None

        with ThreadPoolExecutor(max_workers=len(c_codes)) as executor:
            futures = {
                executor.submit(check_with_both, c_code): c_code
                for c_code in c_codes
            }

            for future in as_completed(futures):
                result = future.result()
                if result and result.is_violation:
                    logger.warning(f"[Excel词表检测-并行+分词] {result.c_code} 检测到违规: {result.keyword}")
                    return result

        logger.info(f"[Excel词表检测-并行+分词] 文本通过检测: {text[:50]}...")
        return ViolationResult(
            is_violation=False,
            c_code="",
            risk_subclass="",
            keyword=""
        )


_parallel_keyword_detector = None


def get_parallel_keyword_detector() -> ParallelKeywordDetector:
    """获取Excel词表检测器单例"""
    global _parallel_keyword_detector
    if _parallel_keyword_detector is None:
        _parallel_keyword_detector = ParallelKeywordDetector()
    return _parallel_keyword_detector
